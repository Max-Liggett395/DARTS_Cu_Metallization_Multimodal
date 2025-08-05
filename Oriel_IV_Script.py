# -*- coding: utf-8 -*-
"""
Created on Fri Jul 11 09:25:00 2025

@author: ethan
"""

# Importing required libraries
import os  # For file and directory operations (listing files, joining paths, etc.)
import pandas as pd  # For handling tabular data (reading, processing, and manipulating data)
from openpyxl import Workbook  # For creating and saving Excel files (.xlsx format)
from openpyxl.utils.dataframe import dataframe_to_rows  # To convert pandas DataFrames into Excel-readable row format
import matplotlib.pyplot as plt # To plot I-V data and be able to frame data
from matplotlib import cm # To do different color gradients for Overlay plot

# Function to extract summary data from a single .txt file
def read_and_process_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Find the line where headers start (line that contains "Voc V")
    start_index = next((i for i, line in enumerate(lines) if "Voc V" in line), None)
    if start_index is None:
        print(f"Error: 'Voc V' line not found in file: {file_path}")
        return None

    # Extract the header and the corresponding data row
    header_line = lines[start_index].strip().split('\t')
    data_line = lines[start_index + 1].strip().split('\t')

    # List of the columns to extract
    relevant_columns = ["Voc V", "Isc A", "Jsc mA/cm2", "Imax A", "Vmax V", 
                        "Pmax mW", "Fill Factor", "Efficiency", "R at Voc", "R at Isc", "Power W"]

    # Find the indices of the required columns
    indices = [header_line.index(col) if col in header_line else None for col in relevant_columns]

    # Check if any required column was not found
    if None in indices:
        print(f"Error: Not all required columns found in file: {file_path}")
        return None

    # Extract only the needed values using the column indices
    extracted_values = [data_line[i] for i in indices]

    # Use the filename (without extension) as the sample name
    sample_name = os.path.splitext(os.path.basename(file_path))[0]

    # Return a dictionary with sample name and extracted values
    return {"Sample": sample_name, **dict(zip(relevant_columns, extracted_values))}

# Function to extract IV curve data from the same file
def read_iv_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Find the start of IV data (line that contains "Vmeas")
    start_index = next((i for i, line in enumerate(lines) if "Vmeas" in line), None)
    if start_index is None:
        print(f"Error: 'Vmeas' line not found in file: {file_path}")
        return None

    # Extract all lines after the header
    iv_lines = lines[start_index + 1:]

    # Split each line into Vmeas and Imeas values
    iv_data = [line.strip().split('\t') for line in iv_lines if line.strip()]
    if not iv_data:
        print(f"Error: No IV data found in file: {file_path}")
        return None

    # Convert to DataFrame for easier manipulation
    df = pd.DataFrame(iv_data, columns=["Vmeas", "Imeas"])
    return df

# Function to process all .txt files in a directory
def process_directory(directory_path):
    raw_data_list = []  # List of summary data from each file
    iv_data_dict = {}   # Dictionary of IV DataFrames keyed by sample name

    for file_name in os.listdir(directory_path):
        # Skip files that are not .txt
        if not file_name.lower().endswith('.txt'):
            continue

        file_path = os.path.join(directory_path, file_name)
        # The line below is an optional line used for debugging to ensure all .txt files are being processed
        #print(f"Processing file: {file_path}")

        # Read summary data
        data_values = read_and_process_data(file_path)
        if data_values is not None:
            raw_data_list.append(data_values)

        # Read IV data
        iv_df = read_iv_data(file_path)
        if iv_df is not None:
            sample_name = os.path.splitext(file_name)[0]
            iv_df.insert(0, 'Sample', sample_name)  # Add sample name as a column
            iv_data_dict[sample_name] = iv_df

    return raw_data_list, iv_data_dict

# Function to write the extracted data to an Excel file
def write_data_to_excel(raw_data, iv_data, output_file, outlier_df=None):
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Extracted Raw Data"

    # Write summary data to first sheet
    if raw_data:
        df_raw = pd.DataFrame(raw_data)
        for r in dataframe_to_rows(df_raw, index=False, header=True):
            ws_raw.append(r)
    else:
        ws_raw.append(["No raw data found."])

    # Add second sheet for IV data
    ws_iv = wb.create_sheet(title="IV")

    if iv_data:
        iv_matrices = []
        max_rows = 0  # Track the longest IV data length

        for sample, df in iv_data.items():
            # Create matrix layout for each sample's IV data
            mat = [ [sample, ''] ]  # First row: sample name
            mat.append(['Vmeas', 'Imeas'])  # Second row: column headers
            mat.extend(df.loc[:, ["Vmeas", "Imeas"]].values.tolist())  # IV data rows
            iv_matrices.append(mat)
            if len(mat) > max_rows:
                max_rows = len(mat)

        # Pad each matrix to the max number of rows with blank rows
        for i in range(len(iv_matrices)):
            while len(iv_matrices[i]) < max_rows:
                iv_matrices[i].append(['', ''])

        # Combine matrices side-by-side with a blank column between samples
        combined = []
        for row_idx in range(max_rows):
            combined_row = []
            for mat in iv_matrices:
                combined_row.extend(mat[row_idx])
                combined_row.append('')  # Blank column
            combined.append(combined_row[:-1])  # Remove last blank column

        # Write combined IV data to sheet
        for row in combined:
            ws_iv.append(row)
    else:
        ws_iv.append(["No IV data found."])

# Add third sheet for statistical data
    ws_stats = wb.create_sheet(title="Statistics")
    stats_df = extract_statistics(raw_data)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
        
        # Add fourth sheet for outliers
    ws_outliers = wb.create_sheet(title="Outliers")
    if outlier_df is not None and not outlier_df.empty:
        for r in dataframe_to_rows(outlier_df, index=False, header=True):
            ws_outliers.append(r)
    else:
        ws_outliers.append(["No outliers detected."])
        
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved to: {os.path.abspath(output_file)}")

def plot_iv_curves(iv_data_dict, output_dir):
    # Create output folders
    plot_dir = os.path.join(output_dir, "Individual_IV_Plots")
    cluster_dir = os.path.join(output_dir, "Cluster_IV_Plots")
    os.makedirs(plot_dir, exist_ok=True)
    os.makedirs(cluster_dir, exist_ok=True)

    # Remove 'dark_IV' from the dataset
    filtered_iv_data = {k: v for k, v in iv_data_dict.items() if 'dark_IV' not in k}

    # Setup colormap
    colormap = cm.get_cmap("viridis")
    num_curves = len(filtered_iv_data)
    color_list = [colormap(i / max(1, num_curves - 1)) for i in range(num_curves)]

    # Prepare for main overlay plot
    overlay_fig, overlay_ax = plt.subplots(figsize=(8, 6))

    # Cluster containers
    cluster_1 = {}  # -0.02 to 0.035
    cluster_2 = {}  # 0.036 to 0.049
    cluster_3 = {}  # 0.05 to 0.075

    # Loop through all samples
    for i, (sample, df) in enumerate(filtered_iv_data.items()):
        v = df["Vmeas"].astype(float).values
        i_val = df["Imeas"].astype(float).values
        color = color_list[i]


        # Plot individual IV curve
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(v, i_val, color=color, label=sample)
        ax.set_title(f"I-V Curve: {sample}")
        ax.set_xlabel("Voltage (V)")
        ax.set_ylabel("Current (A)")
        ax.grid(True)
        fig.tight_layout()

        individual_path = os.path.join(plot_dir, f"{sample}_IV_Curve.png")
        fig.savefig(individual_path)
        plt.close(fig)
        print(f"Saved individual plot for {sample} to: {individual_path}")

        # Shade curves on overlay
        shade_val = 0.3 + 0.7 * (i / max(1, num_curves - 1))
        overlay_ax.plot(v, i_val, color=color, label=sample, alpha=shade_val)

        # Determine max current to classify into cluster
        max_current = max(i_val)
        if -0.02 <= max_current <= 0.035:
            cluster_1[sample] = (v, i_val)
        elif 0.036 <= max_current <= 0.049:
            cluster_2[sample] = (v, i_val)
        elif 0.05 <= max_current <= 0.075:
            cluster_3[sample] = (v, i_val)

    # Finalize overlay plot
    overlay_ax.set_title("Overlay of IV Curves")
    overlay_ax.set_xlabel("Voltage (V)")
    overlay_ax.set_ylabel("Current (A)")
    overlay_ax.set_xlim(-0.02, 0.85)
    overlay_ax.set_ylim(-0.02, 0.075)
    overlay_ax.grid(True)
    overlay_ax.legend(fontsize='small', loc='upper left', bbox_to_anchor=(1.05, 1))
    plt.subplots_adjust(right=0.75)

    overlay_path = os.path.join(output_dir, "All_IV_Curves_Overlay.png")
    overlay_fig.savefig(overlay_path, bbox_inches='tight')
    plt.close(overlay_fig)
    print(f"Overlay plot saved to: {overlay_path}")

    # Function to generate cluster plots
    def plot_cluster(cluster_data, name):
        if not cluster_data:
            print(f"No data for cluster: {name}")
            return

        fig, ax = plt.subplots(figsize=(8, 6))
        for i, (sample, (v, i_val)) in enumerate(cluster_data.items()):
            color = colormap(i / max(1, len(cluster_data) - 1))
            ax.plot(v, i_val, color=color, label=sample)
        ax.set_title(f"Cluster Plot: {name}")
        ax.set_xlabel("Voltage (V)")
        ax.set_ylabel("Current (A)")
        ax.set_xlim(-0.02, 0.85)
        ax.set_ylim(-0.02, 0.075)
        ax.grid(True)
        ax.legend(fontsize='small', loc='upper left', bbox_to_anchor=(1.05, 1))
        plt.subplots_adjust(right=0.75)
        path = os.path.join(cluster_dir, f"{name}_Cluster.png")
        fig.savefig(path, bbox_inches='tight')
        plt.close(fig)
        print(f"{name} cluster plot saved to: {path}")

    # Plot each cluster
    plot_cluster(cluster_1, "Low_Current_Cluster_-0.02_to_0.035")
    plot_cluster(cluster_2, "Mid_Current_Cluster_0.036_to_0.049")
    plot_cluster(cluster_3, "High_Current_Cluster_0.05_to_0.075")



def extract_statistics(raw_data):
    # Convert raw_data list of dicts to DataFrame
    df = pd.DataFrame(raw_data)
    
    # List of performance parameters you want stats for
    params = ["Voc V", "Jsc mA/cm2", "Fill Factor"]
    
    # Filter out parameters not present in df
    params = [p for p in params if p in df.columns]

    # Convert columns to numeric, coerce errors to NaN
    for p in params:
        df[p] = pd.to_numeric(df[p], errors='coerce')

    stats_dict = {}
    for p in params:
        col = df[p].dropna()
        if col.empty:
            stats_dict[p] = {"Mean": None, "Median": None, "Mode": None, "Std Dev": None}
            continue
        
        # pandas mode() returns Series of modes, take the first one if exists
        mode_series = col.mode()
        mode_val = mode_series.iloc[0] if not mode_series.empty else None

        stats_dict[p] = {
            "Mean": col.mean(),
            "Median": col.median(),
            "Mode": mode_val,
            "Standard Deviation": col.std()
        }

    stats_df = pd.DataFrame(stats_dict).T  # transpose
    stats_df.index.name = 'Parameter'
    stats_df.reset_index(inplace=True)

    return stats_df

def detect_outliers(raw_data, params=["Pmax mW", "Isc A", "Voc V"], z_thresh=2.5):
    df = pd.DataFrame(raw_data)
    outlier_report = []

    # Only use columns that exist
    params = [p for p in params if p in df.columns]

    # Convert relevant columns to numeric
    for p in params:
        df[p] = pd.to_numeric(df[p], errors='coerce')

    for p in params:
        col = df[p]
        mean = col.mean()
        std = col.std()
        if std == 0 or pd.isna(std):
            continue  # Avoid division by zero

        z_scores = (col - mean) / std

        for idx, z in z_scores.items():
            if pd.isna(z):
                continue
            if abs(z) > z_thresh:
                outlier_report.append({
                    "Sample": df.loc[idx, "Sample"],
                    "Parameter": p,
                    "Value": df.loc[idx, p],
                    "Z-score": round(z, 2)
                })

    return pd.DataFrame(outlier_report)

def print_outlier_summary(outlier_df):
    if outlier_df.empty:
        print("No outliers detected.")
        return

    # Group by Sample and list parameters responsible
    grouped = outlier_df.groupby("Sample")

    print("\nDetected Outliers:")
    for sample, group in grouped:
        params = group["Parameter"].tolist()
        z_scores = group["Z-score"].tolist()
        vals = group["Value"].tolist()
        print(f"- Sample '{sample}' is an outlier for parameter(s):")
        for p, v, z in zip(params, vals, z_scores):
            print(f"    {p} = {v} (Z-score: {z})")
    print()  # blank line after all


# Main entry point
if __name__ == "__main__":
    # Set the directory to scan for .txt files
    directory_path = r"C:\Users\ethan\Desktop\Work\Python\OneDrive_1_6-20-2025"

    # Process the files in the directory
    raw_data, iv_data = process_directory(directory_path)

# Detect outliers
    outlier_df = detect_outliers(raw_data)
    print_outlier_summary(outlier_df)
    
    # Set the output Excel file path
    output_file = os.path.join(directory_path, "test_data.xlsx")

    # Write all data to the Excel file
    write_data_to_excel(raw_data, iv_data, output_file, outlier_df=outlier_df)

    
    # Plot and save IV curves (overlay + individual)
    plot_iv_curves(iv_data, directory_path)